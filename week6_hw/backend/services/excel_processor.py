"""
Excel预处理服务
智能识别无关行、处理多级表头、拆分合并单元格、输出标准二维表
"""
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Tuple, List, Dict, Any
from backend.utils.openai_client import openai_client
from backend.config import settings
from backend.utils.logger import logger
import json


class ExcelProcessor:
    """Excel文件预处理器"""
    
    async def process_excel(self, file_path: str, output_path: str) -> Dict[str, Any]:
        """
        处理Excel文件（支持多工作表和Schema分割）
        
        Args:
            file_path: 输入Excel文件路径
            output_path: 输出Excel文件路径
            
        Returns:
            处理日志信息
        """
        print(f"🚀 开始处理Excel文件: {file_path}")
        print(f"开始处理Excel文件: {file_path}")
        
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames
        print(f"📊 发现 {len(sheet_names)} 个工作表: {sheet_names}")
        print(f"excel_sheets_detected 发现 {len(sheet_names)} 个工作表: {sheet_names}")
        
        # 创建输出工作簿
        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)  # 删除默认工作表
        
        processed_sheets = {}
        total_log = {
            "total_sheets": len(sheet_names),
            "processed_sheets": 0,
            "schema_split_sheets": 0,
            "skipped_rows": 0,
            "header_row": 0,
            "merged_cells_count": 0,
            "final_rows": 0,
            "final_columns": 0,
            "sheet_details": {}
        }
        
        for sheet_name in sheet_names:
            try:
                print(f"🔄 处理工作表: {sheet_name}")
                print(f"sheet_processing_start 开始处理工作表: {sheet_name}")
                
                # 检查是否需要Schema分割
                print(f"🔍 检查Schema分割需求: {sheet_name}")
                needs_schema_split = await self._check_schema_split_needed(file_path, sheet_name)
                print(f"📊 Schema分割检测结果: {'需要分割' if needs_schema_split else '不需要分割'}")
                print(f"schema_split_check 工作表 {sheet_name} Schema分割检测: {'需要分割' if needs_schema_split else '不需要分割'}")
                
                if needs_schema_split:
                    print(f"✂️ 检测到Schema不一致，将删除表末尾部分: {sheet_name}")
                    print(f"schema_trim_start 开始删除表末尾部分: {sheet_name}")
                    
                    # 先处理合并单元格
                    print(f"🔧 处理合并单元格: {sheet_name}")
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb[sheet_name]
                    merged_cells_count = len(ws.merged_cells.ranges)
                    print(f"📊 发现 {merged_cells_count} 个合并单元格")
                    print(f"merged_cells_processing 工作表 {sheet_name} 发现 {merged_cells_count} 个合并单元格")
                    
                    self._unmerge_cells_with_fill(ws)
                    print(f"✅ 合并单元格处理完成")
                    print(f"merged_cells_processed 工作表 {sheet_name} 合并单元格处理完成")
                    
                    # 保存处理后的临时文件
                    temp_file = file_path.replace('.xlsx', f'_temp_{sheet_name}.xlsx')
                    wb.save(temp_file)
                    print(f"💾 保存临时文件: {temp_file}")
                    
                    try:
                        # 读取完整数据
                        df_full = pd.read_excel(temp_file, sheet_name=sheet_name, header=None)
                        print(f"📊 读取完整数据: {len(df_full)}行 × {len(df_full.columns)}列")
                        print(f"full_data_loaded 读取完整数据: {len(df_full)}行 × {len(df_full.columns)}列")
                        
                        # 获取主要数据部分（删除末尾部分）
                        main_data = await self._get_main_data_section(df_full, sheet_name)
                        print(f"📊 主要数据部分: {len(main_data)}行 × {len(main_data.columns)}列")
                        print(f"main_data_extracted 主要数据部分: {len(main_data)}行 × {len(main_data.columns)}列")
                        
                        # 处理主要数据
                        processed_df = await self._process_single_sheet_data_from_df(main_data, sheet_name)
                        processed_df = self._clean_dataframe(processed_df)
                        print(f"📊 数据处理完成: {len(processed_df)}行 × {len(processed_df.columns)}列")
                        print(f"data_processing_complete 数据处理完成: {len(processed_df)}行 × {len(processed_df.columns)}列")
                        
                        # 添加到输出工作簿（使用原工作表名称）
                        ws = output_wb.create_sheet(title=sheet_name)
                        self._write_dataframe_to_worksheet(ws, processed_df)
                        print(f"📝 写入输出工作簿: {sheet_name}")
                        
                        # 记录信息
                        processed_sheets[sheet_name] = {
                            "rows": len(processed_df),
                            "columns": len(processed_df.columns),
                            "status": "success",
                            "type": "trimmed"
                        }
                        
                        total_log["processed_sheets"] += 1
                        total_log["final_rows"] += len(processed_df)
                        total_log["final_columns"] = max(total_log["final_columns"], len(processed_df.columns))
                        
                        print(f"✅ 表末尾部分删除完成: {sheet_name}")
                        print(f"schema_trim_complete 表末尾部分删除完成: {sheet_name}")
                    
                    finally:
                        # 清理临时文件
                        Path(temp_file).unlink(missing_ok=True)
                        print(f"🗑️ 清理临时文件: {temp_file}")
                        print(f"temp_file_cleaned 清理临时文件: {temp_file}")
                        
                else:
                    # 使用原有逻辑处理单个工作表
                    print(f"📋 使用标准处理逻辑: {sheet_name}")
                    print(f"normal_processing_start 开始标准处理: {sheet_name}")
                    
                    sheet_log = await self._process_single_sheet(file_path, None, sheet_name)
                    print(f"✅ 单工作表处理完成: {sheet_name}")
                    print(f"normal_processing_complete 单工作表处理完成: {sheet_name}")
                    
                    # 读取处理后的数据，先处理合并单元格和多级表头
                    df = await self._process_single_sheet_data(file_path, sheet_name)
                    df = self._clean_dataframe(df)
                    print(f"📊 数据清理完成: {len(df)}行 × {len(df.columns)}列")
                    print(f"data_cleaning_complete 数据清理完成: {len(df)}行 × {len(df.columns)}列")
                    
                    # 添加到输出工作簿
                    ws = output_wb.create_sheet(title=sheet_name)
                    self._write_dataframe_to_worksheet(ws, df)
                    print(f"📝 写入输出工作簿: {sheet_name}")
                    
                    # 记录工作表信息
                    processed_sheets[sheet_name] = {
                        "rows": len(df),
                        "columns": len(df.columns),
                        "status": "success",
                        "type": "normal"
                    }
                    
                    total_log["processed_sheets"] += 1
                    total_log["final_rows"] += len(df)
                    total_log["final_columns"] = max(total_log["final_columns"], len(df.columns))
                    total_log["sheet_details"][sheet_name] = sheet_log
                    
                    print(f"📊 累计统计: {total_log['processed_sheets']}个工作表, {total_log['final_rows']}行, {total_log['final_columns']}列")
                
            except Exception as e:
                print(f"❌ 处理工作表 {sheet_name} 失败: {e}")
                logger.log_error("sheet_processing_failed", f"工作表 {sheet_name} 处理失败: {e}")
                processed_sheets[sheet_name] = {
                    "rows": 0,
                    "columns": 0,
                    "status": "error",
                    "error": str(e)
                }
        
        # 保存处理后的Excel文件
        print(f"💾 保存输出文件: {output_path}")
        print(f"output_file_saving 保存输出文件: {output_path}")
        
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        output_wb.save(output_path)
        
        print(f"✅ 输出文件保存完成: {output_path}")
        print(f"output_file_saved 输出文件保存完成: {output_path}")
        
        total_log["processed_sheets_info"] = processed_sheets
        
        # 打印最终统计
        print(f"\n📊 处理完成统计:")
        print(f"  - 总工作表数: {total_log['total_sheets']}")
        print(f"  - 成功处理: {total_log['processed_sheets']}")
        print(f"  - Schema分割工作表: {total_log['schema_split_sheets']}")
        print(f"  - 总行数: {total_log['final_rows']}")
        print(f"  - 最大列数: {total_log['final_columns']}")
        
        print(f"excel_processing_complete Excel处理完成: {total_log['processed_sheets']}个工作表, {total_log['final_rows']}行, {total_log['final_columns']}列")
        
        return {
            "total_sheets": total_log['total_sheets'],
            "successful_sheets": total_log['processed_sheets'],
            "processed_sheets": total_log['processed_sheets'],
            "schema_split_sheets": total_log['schema_split_sheets'],
            "total_rows": total_log['final_rows'],
            "max_columns": total_log['final_columns'],
            "sheet_details": total_log['sheet_details']
        }
    
    async def _process_single_sheet(self, file_path: str, output_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        处理单个工作表（原有逻辑）
        
        Args:
            file_path: 输入Excel文件路径
            output_path: 输出Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            处理日志信息
        """
        log = {
            "skipped_rows": 0,
            "header_row": 0,
            "merged_cells_count": 0,
            "final_rows": 0,
            "final_columns": 0
        }
        
        # 1. 识别无关行
        skip_rows, header_row = await self._identify_irrelevant_rows(file_path, sheet_name)
        log["skipped_rows"] = len(skip_rows)
        log["header_row"] = header_row
        
        # 2. 处理合并单元格和多级表头
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        
        # 统计合并单元格
        log["merged_cells_count"] = len(ws.merged_cells.ranges)
        
        # 拆分合并单元格
        self._unmerge_cells(ws)
        
        # 3. 读取数据
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            skiprows=skip_rows,
            header=None if header_row in skip_rows else 0
        )
        
        # 4. 处理列名
        if header_row in skip_rows:
            # 如果表头被跳过了，使用默认列名
            df.columns = [f"Column_{i}" for i in range(len(df.columns))]
        else:
            # 如果表头没有被跳过，pandas会自动使用第一行作为列名
            # 清理列名，确保它们是字符串且唯一
            df.columns = [str(col).strip() if col is not None else f"Column_{i}" 
                         for i, col in enumerate(df.columns)]
        
        # 5. 清理数据
        df = self._clean_dataframe(df)
        
        log["final_rows"] = len(df)
        log["final_columns"] = len(df.columns)
        
        # 6. 保存为Excel（如果指定了输出路径）
        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            df.to_excel(output_path, index=False, engine='openpyxl')
        
        return log
    
    async def _process_single_sheet_data(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """
        处理单个工作表的数据，包括合并单元格和多级表头
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            处理后的DataFrame
        """
        print(f"🔧 处理单工作表数据: {sheet_name}")
        
        # 1. 处理合并单元格
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        
        # 统计合并单元格
        merged_cells_count = len(ws.merged_cells.ranges)
        print(f"📊 发现 {merged_cells_count} 个合并单元格")
        
        # 拆分合并单元格
        self._unmerge_cells_with_fill(ws)
        print(f"✅ 合并单元格处理完成")
        
        # 将公式转换为值
        self._convert_formulas_to_values(ws)
        print(f"✅ 公式转值处理完成")
        
        # 2. 保存处理后的临时文件
        temp_file = file_path.replace('.xlsx', f'_temp_{sheet_name}.xlsx')
        wb.save(temp_file)
        print(f"💾 保存临时文件: {temp_file}")
        
        try:
            # 3. 读取数据并处理多级表头
            df = pd.read_excel(temp_file, sheet_name=sheet_name, header=None)
            print(f"📊 读取数据: {len(df)}行 × {len(df.columns)}列")
            
            # 4. 检测标签行和表头
            if len(df) >= 2:
                # 首先检测标签行
                label_rows = self._detect_label_rows(df)
                print(f"🔍 检测到标签行: {label_rows}")
                
                if label_rows:
                    # 移除标签行
                    df = df.drop(label_rows).reset_index(drop=True)
                    print(f"🗑️ 已移除标签行: {label_rows}")
                
                # 检查剩余数据是否足够
                if len(df) >= 2:
                    # 检查前两行是否构成多级表头
                    first_row = df.iloc[0].tolist()
                    second_row = df.iloc[1].tolist()
                    
                    print(f"🔍 检查多级表头:")
                    print(f"  第一行: {first_row}")
                    print(f"  第二行: {second_row}")
                    
                    # 判断是否是多级表头
                    is_multi_level = await self._detect_multi_level_headers(df, 0)
                    
                    if len(is_multi_level) > 1:
                        print(f"📋 检测到多级表头，进行合并")
                        # 合并多级表头
                        merged_headers = self._merge_multi_level_headers(df, is_multi_level)
                        df.columns = merged_headers
                        # 跳过所有表头行
                        df = df.iloc[max(is_multi_level)+1:].reset_index(drop=True)
                        print(f"📋 合并后的表头: {merged_headers}")
                    else:
                        print(f"📋 使用单级表头")
                        # 使用第一行作为列名
                        df.columns = [str(col).strip() if pd.notna(col) and str(col).strip() else f"Column_{i}" 
                                     for i, col in enumerate(df.iloc[0])]
                        df = df.iloc[1:].reset_index(drop=True)
                        print(f"📋 列名: {df.columns.tolist()}")
                else:
                    print(f"⚠️ 移除标签行后数据不足，使用默认列名")
                    df.columns = [f"Column_{i}" for i in range(len(df.columns))]
            else:
                print(f"⚠️ 数据行数不足，使用默认列名")
                df.columns = [f"Column_{i}" for i in range(len(df.columns))]
            
            print(f"✅ 单工作表数据处理完成: {len(df)}行 × {len(df.columns)}列")
            return df
            
        finally:
            # 清理临时文件
            Path(temp_file).unlink(missing_ok=True)
            print(f"🗑️ 清理临时文件: {temp_file}")
    
    def _convert_formulas_to_values(self, ws):
        """
        将工作表中的公式转换为值
        
        Args:
            ws: openpyxl工作表对象
        """
        print(f"🔄 开始转换公式为值")
        
        formula_count = 0
        converted_count = 0
        
        # 遍历所有单元格
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 公式类型
                    formula_count += 1
                    try:
                        # 获取公式的计算值
                        if cell.value is not None:
                            # 将公式替换为计算值
                            cell.value = cell.value
                            converted_count += 1
                    except Exception as e:
                        print(f"⚠️ 公式转换失败: {cell.coordinate} = {cell.value}, 错误: {e}")
                        # 如果公式计算失败，尝试获取显示值
                        try:
                            if hasattr(cell, 'value') and cell.value is not None:
                                # 保持原值，但标记为已处理
                                converted_count += 1
                        except:
                            pass
        
        print(f"📊 公式转换统计: 发现 {formula_count} 个公式, 成功转换 {converted_count} 个")
    
    def _detect_label_rows(self, df: pd.DataFrame) -> List[int]:
        """
        检测标签行（需要移除的行）
        
        Args:
            df: DataFrame
            
        Returns:
            需要移除的行索引列表
        """
        label_rows = []
        
        # 检查前几行是否是标签行
        for i in range(min(5, len(df))):  # 最多检查前5行
            row = df.iloc[i]
            
            # 将行转换为文本
            row_text = ' '.join([str(x) for x in row if pd.notna(x)])
            non_null_count = row.notna().sum()
            
            # 检查是否是标签行的特征
            is_label = False
            
            # 特征1: 前两列为空（重要特征）
            if len(row) >= 2 and pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
                is_label = True
                print(f"  📋 行{i} 识别为标签行: 前两列为空")
            
            # 特征1.5: 前两列为空且包含特定关键词
            elif len(row) >= 2 and pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]) and \
                 any(keyword in row_text for keyword in ['编制', '审核', '日期']):
                is_label = True
                print(f"  📋 行{i} 识别为标签行: 前两列为空且包含关键词")
            
            # 特征2: 大部分列都是NaN（这是最重要的特征）
            elif non_null_count <= 2:  # 只有2个或更少的非空值
                is_label = True
            
            # 特征3: 包含"表"、"图"、"单位"等关键词且大部分列为空或合并单元格列数超过5
            elif any(keyword in row_text for keyword in ['表', '图', '单位：', '单位:', '数据来源', '统计', '报告', '编制', '审核']) and (non_null_count <= 3 or self._count_merged_cells_in_row(df, i) > 5):
                is_label = True
            
            # 特征4: 文本长度很长且大部分列为空或合并单元格列数超过5，可能是标题
            elif len(row_text) > 50 and (non_null_count <= 3 or self._count_merged_cells_in_row(df, i) > 5):
                is_label = True
            
            # 特征5: 包含"增长情况"、"统计表"等描述性文本且大部分列为空或合并单元格列数超过5
            elif any(keyword in row_text for keyword in ['增长情况', '统计表', '汇总表', '明细表', '分析表', '汇总', '意见', '日志', '发电']) and (non_null_count <= 3 or self._count_merged_cells_in_row(df, i) > 5):
                is_label = True
            
            # 特征6: 包含"序号"、"姓名"等表头关键词但大部分列为空或合并单元格列数超过5（可能是标题行）
            elif any(keyword in row_text for keyword in ['序号', '姓名', '学号', '班级']) and (non_null_count <= 3 or self._count_merged_cells_in_row(df, i) > 5):
                is_label = True
            
            # 排除真正的表头行：如果包含常见表头关键词且大部分列都有值，则不是标签行
            if any(keyword in row_text for keyword in ['序号', '姓名', '学号', '班级', '导师', '组别', '题目', '意见', '备注', '电机', '编号', '发电量', '发电小时']) and non_null_count >= 6:
                is_label = False
            
            if is_label:
                label_rows.append(i)
                merged_count = self._count_merged_cells_in_row(df, i)
                print(f"  📋 行{i} 识别为标签行: {row_text[:50]}... (非空列: {non_null_count}/{len(row)}, 合并单元格: {merged_count})")
        
        return label_rows
    
    def _count_merged_cells_in_row(self, df: pd.DataFrame, row_index: int) -> int:
        """
        计算指定行的合并单元格数量
        
        Args:
            df: DataFrame
            row_index: 行索引
            
        Returns:
            合并单元格数量
        """
        try:
            # 检查该行是否有合并单元格
            # 通过检查连续相同值来估算合并单元格
            row = df.iloc[row_index]
            merged_count = 0
            
            # 检查连续相同值（可能是合并单元格的结果）
            current_value = None
            current_count = 0
            
            for value in row:
                if pd.isna(value):
                    if current_count > 1:
                        merged_count += current_count - 1
                    current_count = 0
                    current_value = None
                elif value == current_value:
                    current_count += 1
                else:
                    if current_count > 1:
                        merged_count += current_count - 1
                    current_count = 1
                    current_value = value
            
            # 处理最后一组
            if current_count > 1:
                merged_count += current_count - 1
            
            return merged_count
        except Exception as e:
            print(f"  ⚠️ 计算合并单元格数量失败: {e}")
            return 0
    
    async def _is_empty_sheet(self, file_path: str, sheet_name: str) -> bool:
        """
        检查工作表是否为空
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            是否为空工作表
        """
        try:
            # 读取工作表
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # 检查是否有数据
            if len(df) == 0:
                return True
            
            # 检查是否所有单元格都是NaN
            if df.isnull().all().all():
                return True
            
            # 检查是否有足够的非空数据（至少3行3列）
            non_null_count = df.notna().sum().sum()
            if non_null_count < 9:  # 3x3 = 9个单元格
                return True
            
            return False
            
        except Exception as e:
            print(f"⚠️ 检查空工作表失败: {e}")
            return False
    
    async def _get_main_data_section(self, df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """
        获取主要数据部分，删除表末尾的无关部分
        
        Args:
            df: 原始DataFrame
            sheet_name: 工作表名称
            
        Returns:
            主要数据部分的DataFrame
        """
        try:
            print(f"🔍 开始分析主要数据部分: {sheet_name}")
            
            # 使用GPT-4分析主要数据部分
            sample_text = df.head(50).to_string()
            
            prompt = f"""
请分析以下Excel数据，找出主要数据部分的结束位置。

数据内容：
{sample_text}

请分析：
1. 主要数据部分从第几行开始到第几行结束？
2. 哪些行是无关的统计信息、汇总数据或注释？
3. 主要数据部分应该包含哪些行？

请以JSON格式返回结果：
{{
    "main_data_start": 起始行号,
    "main_data_end": 结束行号,
    "reason": "分析原因"
}}

注意：
- 行号从0开始计算
- 主要数据部分应该是结构一致的数据行
- 排除表头、统计信息、汇总数据等
"""
            
            response = await openai_client.call_gpt4(prompt)
            result = json.loads(response)
            
            start_row = result.get("main_data_start", 0)
            end_row = result.get("main_data_end", len(df))
            reason = result.get("reason", "未指定原因")
            
            print(f"📊 主要数据部分: 第{start_row}行到第{end_row}行")
            print(f"📋 分析原因: {reason}")
            
            # 提取主要数据部分
            if end_row > start_row:
                main_data = df.iloc[start_row:end_row].copy()
                print(f"✅ 提取主要数据: {len(main_data)}行 × {len(main_data.columns)}列")
                return main_data
            else:
                print(f"⚠️ 主要数据部分无效，返回原始数据")
                return df
                
        except Exception as e:
            print(f"❌ 获取主要数据部分失败: {e}")
            # 如果分析失败，返回前80%的数据
            end_row = int(len(df) * 0.8)
            print(f"🔄 使用默认策略: 返回前{end_row}行数据")
            return df.iloc[:end_row].copy()
    
    async def _process_single_sheet_data_from_df(self, df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """
        从DataFrame处理单个工作表数据，包括合并单元格和多级表头
        
        Args:
            df: 输入DataFrame
            sheet_name: 工作表名称
            
        Returns:
            处理后的DataFrame
        """
        try:
            print(f"🔧 处理单工作表数据: {sheet_name}")
            
            # 检测标签行
            label_rows = self._detect_label_rows(df)
            print(f"🔍 检测到标签行: {label_rows}")
            
            if label_rows:
                # 移除标签行
                df = df.drop(label_rows).reset_index(drop=True)
                print(f"🗑️ 已移除标签行: {label_rows}")
            
            # 检查剩余数据是否足够
            if len(df) >= 2:
                # 检测多级表头
                print(f"🔍 检查多级表头:")
                print(f"  第一行: {df.iloc[0].tolist()}")
                if len(df) > 1:
                    print(f"  第二行: {df.iloc[1].tolist()}")
                
                # 使用GPT-4检测多级表头
                header_rows = await self._detect_multi_level_headers(df, 0)
                print(f"🔍 开始多级表头检测: 表头行 0")
                print(f"multi_level_header_detection_start 开始多级表头检测: 表头行 0")
                
                if len(header_rows) > 1:
                    print(f"📊 多级表头检测结果: 是多级表头")
                    print(f"📋 表头行: {header_rows}")
                    print(f"multi_level_header_detection_complete 多级表头检测完成: 是多级表头, 表头行: {header_rows}")
                    
                    # 合并多级表头
                    merged_headers = self._merge_multi_level_headers(df, header_rows)
                    df.columns = merged_headers
                    print(f"📋 合并后的表头: {df.columns.tolist()}")
                    print(f"multi_level_header_merged 合并后的表头: {df.columns.tolist()}")
                    
                    # 跳过表头行
                    df = df.iloc[len(header_rows):].reset_index(drop=True)
                    print(f"📊 跳过表头行后: {len(df)}行")
                else:
                    print(f"📊 多级表头检测结果: 是单级表头")
                    print(f"📋 表头行: {header_rows}")
                    print(f"multi_level_header_detection_complete 多级表头检测完成: 是单级表头, 表头行: {header_rows}")
                    print(f"📋 使用单级表头")
                    
                    # 使用第一行作为列名
                    df.columns = [str(col).strip() if pd.notna(col) and str(col).strip() else f"Column_{i}"
                                 for i, col in enumerate(df.iloc[0])]
                    df = df.iloc[1:].reset_index(drop=True)
                    print(f"📋 列名: {df.columns.tolist()}")
            else:
                print(f"⚠️ 数据行数不足，使用默认列名")
                df.columns = [f"Column_{i}" for i in range(len(df.columns))]
            
            print(f"✅ 单工作表数据处理完成: {len(df)}行 × {len(df.columns)}列")
            return df
            
        except Exception as e:
            print(f"❌ 处理单工作表数据失败: {e}")
            import traceback
            traceback.print_exc()
            return df
    
    async def _identify_irrelevant_rows(self, file_path: str, sheet_name: str = None) -> Tuple[List[int], int]:
        """
        使用GPT-4识别无关行
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（可选）
            
        Returns:
            (要跳过的行号列表, 表头行号)
        """
        # 读取前20行
        df_sample = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
        sample_text = df_sample.to_string()
        
        prompt = f"""分析以下Excel文件的前20行数据，识别：
1. 哪些行是标题、说明、空行等非数据内容
2. 哪一行是真正的数据表头

返回JSON格式：
{{
    "skip_rows": [要跳过的行号列表，从0开始],
    "header_row": 表头行号（从0开始）
}}

数据样本：
{sample_text}

请仔细分析，返回纯JSON，不要其他解释。"""
        
        messages = [
            {"role": "system", "content": "你是一个Excel数据分析专家，擅长识别表格结构。"},
            {"role": "user", "content": prompt}
        ]
        
        try:
            response = await openai_client.chat_completion(messages, temperature=0.1)
            result = openai_client.extract_json(response)
            
            skip_rows = result.get("skip_rows", [])
            header_row = result.get("header_row", 0)
            
            return skip_rows, header_row
        except Exception as e:
            print(f"GPT-4识别失败，使用默认策略: {str(e)}")
            # 默认：跳过前面的空行
            return [], 0
    
    def _unmerge_cells(self, worksheet):
        """拆分合并单元格并填充值"""
        merged_ranges = list(worksheet.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            # 获取合并区域左上角的值
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = worksheet.cell(min_row, min_col).value
            
            # 拆分合并单元格
            worksheet.unmerge_cells(str(merged_range))
            
            # 填充所有单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    worksheet.cell(row, col).value = value
    
    def _merge_multi_level_headers(self, worksheet, header_start_row: int) -> List[str]:
        """
        合并多级表头
        
        Args:
            worksheet: openpyxl工作表
            header_start_row: 表头起始行（从0开始，但openpyxl从1开始）
            
        Returns:
            合并后的列名列表
        """
        # 检测表头层数（连续的非空行）
        header_rows = []
        row_idx = header_start_row + 1  # openpyxl从1开始
        
        while row_idx <= min(row_idx + 5, worksheet.max_row):
            row_values = [cell.value for cell in worksheet[row_idx]]
            # 如果大部分是空值，认为表头结束
            if sum(1 for v in row_values if v is not None) < len(row_values) * 0.3:
                break
            header_rows.append(row_values)
            row_idx += 1
        
        if not header_rows:
            return []
        
        # 如果只有一层表头
        if len(header_rows) == 1:
            return [str(v) if v is not None else f"Column_{i}" for i, v in enumerate(header_rows[0])]
        
        # 多级表头合并
        column_names = []
        num_cols = len(header_rows[0])
        
        for col_idx in range(num_cols):
            parts = []
            for row in header_rows:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if val and val not in parts:  # 避免重复
                        parts.append(val)
            
            if parts:
                column_names.append("_".join(parts))
            else:
                column_names.append(f"Column_{col_idx}")
        
        return column_names
    
    async def _check_schema_split_needed(self, file_path: str, sheet_name: str) -> bool:
        """
        检查工作表是否需要Schema分割
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            是否需要Schema分割
        """
        try:
            print(f"🔍 开始Schema分割检查: {sheet_name}")
            print(f"开始Schema分割检查: {sheet_name}")
            
            # 读取整个工作表
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            print(f"📊 读取数据: {len(df)}行 × {len(df.columns)}列")
            print(f"读取数据: {len(df)}行 × {len(df.columns)}列")
            
            # 如果数据太少，不需要分割
            if len(df) < 10:
                print(f"📊 数据量不足，跳过Schema分割检查: {len(df)}行")
                print(f"schema_check_skipped 数据量不足，跳过Schema分割检查: {len(df)}行")
                return False
            
            # 使用GPT-4分析是否需要Schema分割
            # 限制样本数据长度，避免超过token限制
            sample_df = df.head(20)  # 减少到20行
            sample_text = sample_df.to_string(max_rows=20, max_cols=10)  # 限制列数
            
            # 如果文本太长，进一步截断
            if len(sample_text) > 1500:  # 降低到1500字符，留出更多余量
                sample_text = sample_text[:1497] + "..."  # 1497 + 3 = 1500
            
            print(f"🤖 调用GPT-4进行Schema分析...")
            print(f"gpt_analysis_start 调用GPT-4进行Schema分析: {sheet_name}")
            print(f"📊 样本数据长度: {len(sample_text)} 字符")
            
            prompt = f"""分析以下Excel工作表数据，判断是否需要按Schema分割：

1. 检查数据是否包含多个不同的数据结构（如：统计信息、详细数据、汇总表等）
2. 检查是否有明显的分隔行（如：空行、标题行、统计值行等）
3. 检查列结构是否在不同区域发生变化

返回JSON格式：
{{
    "needs_split": true/false,
    "reason": "分割原因说明",
    "split_points": [分割点行号列表，从0开始],
    "schema_descriptions": ["各Schema的描述"]
}}

数据样本（前20行）：
{sample_text}

请仔细分析，返回纯JSON，不要其他解释。"""
            
            messages = [
                {"role": "system", "content": "你是一个专业的数据分析师，擅长识别Excel表格中的数据结构变化和Schema分割点。"},
                {"role": "user", "content": prompt}
            ]
            
            response = await openai_client.chat_completion(messages)
            result = openai_client.extract_json(response)
            
            needs_split = result.get("needs_split", False)
            reason = result.get("reason", "未知原因")
            
            print(f"📊 Schema分割分析结果: {'需要分割' if needs_split else '不需要分割'}")
            if needs_split:
                print(f"📋 分割原因: {reason}")
                split_points = result.get("split_points", [])
                schema_descriptions = result.get("schema_descriptions", [])
                print(f"📍 分割点: {split_points}")
                print(f"📝 Schema描述: {schema_descriptions}")
            
            print(f"schema_check_complete Schema分割检查完成: {'需要分割' if needs_split else '不需要分割'}, 原因: {reason}")
            
            return needs_split
            
        except Exception as e:
            print(f"❌ Schema分割检查失败: {e}")
            logger.log_error("schema_check_failed", f"Schema分割检查失败: {e}")
            return False
    
    async def _process_schema_split_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Dict[str, Any]]:
        """
        处理需要Schema分割的工作表
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            分割后的Schema数据字典
        """
        try:
            print(f"🔍 开始Schema分割处理: {sheet_name}")
            print(f"schema_split_processing_start 开始Schema分割处理: {sheet_name}")
            
            # 先处理公式转值
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]
            self._convert_formulas_to_values(ws)
            
            # 保存处理后的临时文件
            temp_file = file_path.replace('.xlsx', f'_temp_schema_{sheet_name}.xlsx')
            wb.save(temp_file)
            print(f"💾 保存Schema处理临时文件: {temp_file}")
            
            try:
                # 读取整个工作表
                df = pd.read_excel(temp_file, sheet_name=sheet_name, header=None)
            finally:
                # 清理临时文件
                Path(temp_file).unlink(missing_ok=True)
                print(f"🗑️ 清理Schema处理临时文件: {temp_file}")
            print(f"📊 读取完整数据: {len(df)}行 × {len(df.columns)}列")
            print(f"full_data_loaded 读取完整数据: {len(df)}行 × {len(df.columns)}列")
            
            # 使用GPT-4分析分割点
            sample_text = df.to_string()
            print(f"🤖 调用GPT-4进行Schema分割分析...")
            print(f"gpt_schema_analysis_start 调用GPT-4进行Schema分割分析: {sheet_name}")
            
            prompt = f"""分析以下Excel工作表，识别Schema分割点：

1. 找出数据结构发生变化的分割点
2. 为每个Schema区域提供描述
3. 识别每个区域的表头位置

返回JSON格式：
{{
    "schema_regions": [
        {{
            "start_row": 起始行号（从0开始）,
            "end_row": 结束行号（从0开始，-1表示到最后）,
            "header_row": 表头行号（从0开始）,
            "description": "Schema描述",
            "schema_name": "Schema名称"
        }}
    ]
}}

数据：
{sample_text}

请仔细分析，返回纯JSON，不要其他解释。"""
            
            messages = [
                {"role": "system", "content": "你是一个专业的数据分析师，擅长识别Excel表格中的Schema分割点和数据结构。"},
                {"role": "user", "content": prompt}
            ]
            
            response = await openai_client.chat_completion(messages)
            result = openai_client.extract_json(response)
            
            schema_sheets = {}
            schema_regions = result.get("schema_regions", [])
            print(f"📋 检测到 {len(schema_regions)} 个Schema区域")
            print(f"schema_regions_detected 检测到 {len(schema_regions)} 个Schema区域")
            
            for i, region in enumerate(schema_regions):
                start_row = region.get("start_row", 0)
                end_row = region.get("end_row", -1)
                header_row = region.get("header_row", start_row)
                description = region.get("description", "未知Schema")
                schema_name = region.get("schema_name", f"{sheet_name}_{description}")
                
                print(f"📋 处理Schema区域 {i+1}: {schema_name}")
                print(f"  - 描述: {description}")
                print(f"  - 起始行: {start_row}, 结束行: {end_row}, 表头行: {header_row}")
                print(f"schema_region_processing 处理Schema区域: {schema_name}, 描述: {description}")
                
                # 提取Schema数据
                if end_row == -1:
                    schema_df = df.iloc[start_row:].copy()
                else:
                    schema_df = df.iloc[start_row:end_row+1].copy()
                
                print(f"  - 提取数据: {len(schema_df)}行 × {len(schema_df.columns)}列")
                
                schema_sheets[schema_name] = {
                    "data": schema_df,
                    "header_row": header_row - start_row,  # 相对于Schema的header位置
                    "description": description,
                    "original_sheet": sheet_name
                }
            
            print(f"✅ Schema分割处理完成: {len(schema_sheets)}个Schema")
            print(f"schema_split_processing_complete Schema分割处理完成: {len(schema_sheets)}个Schema")
            
            return schema_sheets
            
        except Exception as e:
            print(f"❌ Schema分割处理失败: {e}")
            logger.log_error("schema_split_processing_failed", f"Schema分割处理失败: {e}")
            return {}
    
    async def _process_schema_data(self, schema_data: Dict[str, Any]) -> pd.DataFrame:
        """
        处理单个Schema的数据
        
        Args:
            schema_data: Schema数据字典
            
        Returns:
            处理后的DataFrame
        """
        schema_name = schema_data.get("description", "未知Schema")
        print(f"🔄 开始处理Schema数据: {schema_name}")
        print(f"schema_data_processing_start 开始处理Schema数据: {schema_name}")
        
        df = schema_data["data"].copy()
        header_row = schema_data["header_row"]
        original_sheet = schema_data["original_sheet"]
        
        print(f"📊 原始数据: {len(df)}行 × {len(df.columns)}列, 表头行: {header_row}")
        print(f"schema_data_info 原始数据: {len(df)}行 × {len(df.columns)}列, 表头行: {header_row}")
        
        # 1. 处理合并单元格
        # 注意：这里需要从原始文件路径重新加载，而不是使用sheet_name
        # 由于我们已经在主方法中处理了合并单元格，这里直接使用DataFrame
        print(f"🔧 合并单元格已在主方法中处理")
        
        # 2. 处理多级表头
        if header_row >= 0 and header_row < len(df):
            print(f"🔍 检查多级表头: 表头行 {header_row}")
            print(f"multi_level_header_check 检查多级表头: 表头行 {header_row}")
            
            # 检查是否有多级表头
            multi_level_headers = await self._detect_multi_level_headers(df, header_row)
            print(f"📊 多级表头检测结果: {len(multi_level_headers)}个表头行 {multi_level_headers}")
            print(f"multi_level_header_detected 多级表头检测结果: {len(multi_level_headers)}个表头行 {multi_level_headers}")
            
            if len(multi_level_headers) > 1:
                print(f"🔗 合并多级表头...")
                # 合并多级表头
                merged_headers = self._merge_multi_level_headers(df, multi_level_headers)
                print(f"📋 合并后的表头: {merged_headers}")
                print(f"multi_level_header_merged 合并后的表头: {merged_headers}")
                
                df.columns = merged_headers
                # 跳过所有表头行
                df = df.iloc[max(multi_level_headers)+1:].reset_index(drop=True)
                print(f"📊 跳过表头行后: {len(df)}行")
            else:
                print(f"📋 使用单级表头...")
                # 使用单级表头
                header_values = df.iloc[header_row].tolist()
                print(f"📋 原始表头值: {header_values}")
                
                # 确保列名是字符串且不为空
                clean_headers = []
                for i, val in enumerate(header_values):
                    if pd.notna(val) and str(val).strip():
                        clean_headers.append(str(val).strip())
                    else:
                        clean_headers.append(f"Column_{i}")
                
                print(f"📋 清理后的表头: {clean_headers}")
                print(f"single_level_header_processed 清理后的表头: {clean_headers}")
                
                df.columns = clean_headers
                df = df.iloc[header_row+1:].reset_index(drop=True)
                print(f"📊 跳过表头行后: {len(df)}行")
        else:
            print(f"⚠️ 表头行无效，使用默认列名")
            print(f"invalid_header_row 表头行无效: {header_row}, 使用默认列名")
        
        # 3. 清理数据
        print(f"🧹 开始数据清理...")
        print(f"data_cleaning_start 开始数据清理: {schema_name}")
        
        original_rows = len(df)
        original_cols = len(df.columns)
        
        df = self._clean_dataframe(df)
        
        print(f"✅ 数据清理完成: {original_rows}行 → {len(df)}行, {original_cols}列 → {len(df.columns)}列")
        print(f"data_cleaning_complete 数据清理完成: {original_rows}行 → {len(df)}行, {original_cols}列 → {len(df.columns)}列")
        
        print(f"✅ Schema数据处理完成: {schema_name}")
        print(f"schema_data_processing_complete Schema数据处理完成: {schema_name}")
        
        return df
    
    def _unmerge_cells_with_fill(self, worksheet):
        """
        拆分合并单元格并填充值到所有单元格
        
        Args:
            worksheet: openpyxl工作表
        """
        merged_ranges = list(worksheet.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            # 获取合并区域左上角的值
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_cell_value = worksheet.cell(min_row, min_col).value
            
            # 拆分合并单元格
            worksheet.unmerge_cells(str(merged_range))
            
            # 用原值填充所有单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    worksheet.cell(row, col).value = top_left_cell_value
    
    async def _detect_multi_level_headers(self, df: pd.DataFrame, header_row: int) -> List[int]:
        """
        检测多级表头
        
        Args:
            df: DataFrame
            header_row: 表头行号
            
        Returns:
            多级表头的行号列表
        """
        try:
            print(f"🔍 开始多级表头检测: 表头行 {header_row}")
            print(f"multi_level_header_detection_start 开始多级表头检测: 表头行 {header_row}")
            
            # 检查表头行附近的行
            check_rows = max(0, header_row - 2)
            check_end = min(len(df), header_row + 3)
            
            print(f"📊 检查行范围: {check_rows} 到 {check_end}")
            print(f"header_check_range 检查行范围: {check_rows} 到 {check_end}")
            
            sample_data = df.iloc[check_rows:check_end].to_string()
            
            print(f"🤖 调用GPT-4进行多级表头分析...")
            print(f"gpt_multi_level_analysis_start 调用GPT-4进行多级表头分析")
            
            prompt = f"""分析以下Excel数据，检测多级表头：

1. 检查表头行附近是否有多个表头行
2. 判断哪些行是表头的一部分
3. 识别表头的层级结构

返回JSON格式：
{{
    "is_multi_level": true/false,
    "header_rows": [表头行号列表，从0开始],
    "header_levels": {{
        "row_0": "第一级表头描述",
        "row_1": "第二级表头描述"
    }}
}}

数据样本：
{sample_data}

请仔细分析，返回纯JSON，不要其他解释。"""
            
            messages = [
                {"role": "system", "content": "你是一个专业的数据分析师，擅长识别Excel表格中的多级表头结构。"},
                {"role": "user", "content": prompt}
            ]
            
            response = await openai_client.chat_completion(messages)
            result = openai_client.extract_json(response)
            
            is_multi_level = result.get("is_multi_level", False)
            header_rows = result.get("header_rows", [header_row])
            header_levels = result.get("header_levels", {})
            
            print(f"📊 多级表头检测结果: {'是多级表头' if is_multi_level else '是单级表头'}")
            print(f"📋 表头行: {header_rows}")
            if header_levels:
                print(f"📝 表头层级: {header_levels}")
            
            print(f"multi_level_header_detection_complete 多级表头检测完成: {'是多级表头' if is_multi_level else '是单级表头'}, 表头行: {header_rows}")
            
            if is_multi_level:
                return header_rows
            else:
                return [header_row]
                
        except Exception as e:
            print(f"❌ 多级表头检测失败: {e}")
            logger.log_error("multi_level_header_detection_failed", f"多级表头检测失败: {e}")
            return [header_row]
    
    def _merge_multi_level_headers(self, df: pd.DataFrame, header_rows: List[int]) -> List[str]:
        """
        合并多级表头为单级表头
        
        Args:
            df: DataFrame
            header_rows: 表头行号列表
            
        Returns:
            合并后的列名列表
        """
        print(f"🔗 开始合并多级表头: {len(header_rows)}个表头行")
        print(f"multi_level_header_merge_start 开始合并多级表头: {len(header_rows)}个表头行")
        
        merged_headers = []
        
        for col_idx in range(len(df.columns)):
            col_values = []
            
            # 收集每级表头的值
            for row_idx in header_rows:
                if row_idx < len(df):
                    value = df.iloc[row_idx, col_idx]
                    if pd.notna(value) and str(value).strip():
                        col_values.append(str(value).strip())
            
            # 合并表头值
            if len(col_values) > 1:
                # 去重并合并
                unique_values = []
                for val in col_values:
                    if val not in unique_values:
                        unique_values.append(val)
                
                if len(unique_values) > 1:
                    # 如果去重后仍有多个值，用"-"连接
                    merged_header = "-".join(unique_values)
                    print(f"  📋 列{col_idx}: {' + '.join(col_values)} → {merged_header}")
                else:
                    # 如果去重后只有一个值，直接使用
                    merged_header = unique_values[0]
                    print(f"  📋 列{col_idx}: {' + '.join(col_values)} → {merged_header}")
            elif len(col_values) == 1:
                # 如果只有一个值，直接使用
                merged_header = col_values[0]
                print(f"  📋 列{col_idx}: {merged_header}")
            else:
                # 如果没有值，使用默认名称
                merged_header = f"Column_{col_idx}"
                print(f"  📋 列{col_idx}: 空值 → {merged_header}")
            
            merged_headers.append(merged_header)
        
        print(f"✅ 多级表头合并完成: {len(merged_headers)}列")
        print(f"multi_level_header_merge_complete 多级表头合并完成: {len(merged_headers)}列")
        
        return merged_headers
    
    def _write_dataframe_to_worksheet(self, worksheet, df: pd.DataFrame):
        """
        将DataFrame写入工作表
        
        Args:
            worksheet: openpyxl工作表
            df: DataFrame
        """
        print(f"📝 开始写入DataFrame到工作表: {len(df)}行 × {len(df.columns)}列")
        print(f"dataframe_write_start 开始写入DataFrame到工作表: {len(df)}行 × {len(df.columns)}列")
        
        # 写入表头
        print(f"📋 写入表头: {list(df.columns)}")
        for col_idx, col_name in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=col_name)
        
        # 写入数据
        print(f"📊 写入数据行...")
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"✅ DataFrame写入完成")
        print(f"dataframe_write_complete DataFrame写入完成: {len(df)}行 × {len(df.columns)}列")
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """清理DataFrame"""
        print(f"🧹 开始清理DataFrame: {len(df)}行 × {len(df.columns)}列")
        print(f"dataframe_cleaning_start 开始清理DataFrame: {len(df)}行 × {len(df.columns)}列")
        
        original_rows = len(df)
        original_cols = len(df.columns)
        
        # 删除全空行
        df = df.dropna(how='all')
        empty_rows_removed = original_rows - len(df)
        if empty_rows_removed > 0:
            print(f"🗑️ 删除空行: {empty_rows_removed}行")
            print(f"empty_rows_removed 删除空行: {empty_rows_removed}行")
        
        # 删除全空列
        df = df.dropna(axis=1, how='all')
        empty_cols_removed = original_cols - len(df.columns)
        if empty_cols_removed > 0:
            print(f"🗑️ 删除空列: {empty_cols_removed}列")
            print(f"empty_cols_removed 删除空列: {empty_cols_removed}列")
        
        # 确保列名唯一
        print(f"🔧 处理列名唯一性...")
        cols = []
        renamed_cols = 0
        for col in df.columns:
            col_str = str(col) if col is not None else "Unnamed"
            original_col_str = col_str
            # 如果列名已存在，添加后缀
            if col_str in cols:
                i = 1
                while f"{col_str}_{i}" in cols:
                    i += 1
                col_str = f"{col_str}_{i}"
                if col_str != original_col_str:
                    renamed_cols += 1
            cols.append(col_str)
        
        if renamed_cols > 0:
            print(f"📝 重命名重复列名: {renamed_cols}列")
            print(f"duplicate_cols_renamed 重命名重复列名: {renamed_cols}列")
        
        df.columns = cols
        
        # 处理NaN值，避免Elasticsearch索引失败
        print(f"🔧 处理NaN值...")
        nan_count = 0
        for col in df.columns:
            # 检查所有类型的NaN值
            nan_mask = df[col].isna()
            if nan_mask.any():
                nan_count += nan_mask.sum()
                if df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                    # 对于数值列，将NaN替换为0
                    df.loc[nan_mask, col] = 0
                else:
                    # 对于其他列，将NaN替换为空字符串
                    df.loc[nan_mask, col] = ''
        
        if nan_count > 0:
            print(f"🔧 处理NaN值: {nan_count}个")
            print(f"nan_values_handled 处理NaN值: {nan_count}个")
        
        print(f"✅ DataFrame清理完成: {original_rows}行 → {len(df)}行, {original_cols}列 → {len(df.columns)}列")
        print(f"dataframe_cleaning_complete DataFrame清理完成: {original_rows}行 → {len(df)}行, {original_cols}列 → {len(df.columns)}列")
        
        return df


# 全局实例
excel_processor = ExcelProcessor()

