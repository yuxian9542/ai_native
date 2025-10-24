#!/usr/bin/env python3
"""
测试增强的Excel处理器功能
包括Schema分割、合并单元格处理、多级表头合并和详细日志
"""
import sys
import asyncio
from pathlib import Path
import pandas as pd

# 添加项目根目录到Python路径
sys.path.insert(0, str(Path(__file__).parent.parent))

async def test_enhanced_excel_processor():
    """测试增强的Excel处理器功能"""
    print("🧪 测试增强的Excel处理器功能")
    print("=" * 60)
    
    from backend.services.excel_processor import excel_processor
    
    # 测试文件路径
    test_file = Path(__file__).parent.parent / "backend" / "data" / "original" / "发电日志.xlsx"
    output_file = Path(__file__).parent / "data" / "processed" / "发电日志_enhanced_test.xlsx"
    
    # 确保输出目录存在
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    if not test_file.exists():
        print(f"❌ 测试文件不存在: {test_file}")
        print("请确保发电日志.xlsx文件存在于backend/data/original/目录中")
        return
    
    try:
        print(f"📁 测试文件: {test_file}")
        print(f"📁 输出文件: {output_file}")
        print("\n" + "="*60)
        print("🚀 开始处理，观察详细日志输出...")
        print("="*60)
        
        # 处理Excel文件
        result = await excel_processor.process_excel(str(test_file), str(output_file))
        
        print("\n" + "="*60)
        print("📊 处理结果总结:")
        print("="*60)
        print(f"  - 总工作表数: {result.get('total_sheets', 0)}")
        print(f"  - 成功处理: {result.get('processed_sheets', 0)}")
        print(f"  - Schema分割工作表: {result.get('schema_split_sheets', 0)}")
        print(f"  - 总行数: {result.get('final_rows', 0)}")
        print(f"  - 最大列数: {result.get('final_columns', 0)}")
        
        # 显示每个工作表的详细信息
        if 'processed_sheets_info' in result:
            print(f"\n📋 工作表详情:")
            for sheet_name, info in result['processed_sheets_info'].items():
                status = "✅" if info['status'] == 'success' else "❌"
                sheet_type = info.get('type', 'normal')
                print(f"  {status} {sheet_name} ({sheet_type}): {info['rows']}行 × {info['columns']}列")
                if info['status'] == 'error':
                    print(f"    错误: {info.get('error', '未知错误')}")
        
        # 验证输出文件
        if output_file.exists():
            print(f"\n✅ 输出文件已生成: {output_file}")
            print(f"📊 文件大小: {output_file.stat().st_size / 1024:.1f} KB")
            
            # 读取输出文件验证
            excel_file = pd.ExcelFile(str(output_file))
            print(f"📊 输出文件包含 {len(excel_file.sheet_names)} 个工作表:")
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(str(output_file), sheet_name=sheet_name)
                print(f"\n  📋 工作表: {sheet_name}")
                print(f"    - 行数: {len(df)}")
                print(f"    - 列数: {len(df.columns)}")
                print(f"    - 列名: {list(df.columns)}")
                
                # 显示前几行数据
                if len(df) > 0:
                    print(f"    - 前3行数据:")
                    for i, row in df.head(3).iterrows():
                        print(f"      行{i+1}: {dict(row)}")
        else:
            print("❌ 输出文件未生成")
            
        print(f"\n📝 详细日志已记录到系统日志中")
        print(f"🔍 查看日志文件: backend/logs/ 和 test/logs/ 目录")
        
    except Exception as e:
        print(f"❌ 处理失败: {e}")
        import traceback
        traceback.print_exc()

async def test_schema_split_detection():
    """测试Schema分割检测功能"""
    print("\n🔍 测试Schema分割检测功能")
    print("=" * 40)
    
    from backend.services.excel_processor import excel_processor
    
    # 测试文件
    test_file = Path(__file__).parent.parent / "backend" / "data" / "original" / "发电日志.xlsx"
    
    if not test_file.exists():
        print(f"❌ 测试文件不存在: {test_file}")
        return
    
    try:
        # 测试Schema分割检测
        print(f"🔍 测试Schema分割检测...")
        needs_split = await excel_processor._check_schema_split_needed(str(test_file), "10月")
        print(f"📊 Schema分割检测结果: {'需要分割' if needs_split else '不需要分割'}")
        
        if needs_split:
            # 测试Schema分割处理
            print(f"🔍 测试Schema分割处理...")
            schema_sheets = await excel_processor._process_schema_split_sheet(str(test_file), "10月")
            print(f"📋 检测到 {len(schema_sheets)} 个Schema区域:")
            
            for schema_name, schema_data in schema_sheets.items():
                print(f"  - {schema_name}: {schema_data['description']}")
                print(f"    数据行数: {len(schema_data['data'])}")
                print(f"    表头行: {schema_data['header_row']}")
        
    except Exception as e:
        print(f"❌ Schema分割检测失败: {e}")
        import traceback
        traceback.print_exc()

async def test_multi_level_header_detection():
    """测试多级表头检测功能"""
    print("\n🔍 测试多级表头检测功能")
    print("=" * 40)
    
    from backend.services.excel_processor import excel_processor
    
    # 创建测试数据
    test_data = pd.DataFrame({
        'A': ['一级表头1', '二级表头1', '数据1', '数据2'],
        'B': ['一级表头1', '二级表头2', '数据3', '数据4'],
        'C': ['一级表头2', '二级表头3', '数据5', '数据6']
    })
    
    try:
        # 测试多级表头检测
        print(f"🔍 测试多级表头检测...")
        header_rows = await excel_processor._detect_multi_level_headers(test_data, 1)
        print(f"📊 检测到的表头行: {header_rows}")
        
        if len(header_rows) > 1:
            # 测试表头合并
            print(f"🔗 测试表头合并...")
            merged_headers = excel_processor._merge_multi_level_headers(test_data, header_rows)
            print(f"📋 合并后的表头: {merged_headers}")
        
    except Exception as e:
        print(f"❌ 多级表头检测失败: {e}")
        import traceback
        traceback.print_exc()

async def test_merged_cells_processing():
    """测试合并单元格处理功能"""
    print("\n🔍 测试合并单元格处理功能")
    print("=" * 40)
    
    from backend.services.excel_processor import excel_processor
    import openpyxl
    from pathlib import Path
    
    # 创建测试Excel文件
    test_file = Path(__file__).parent / "data" / "original" / "test_merged_cells.xlsx"
    test_file.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        # 创建包含合并单元格的测试文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试合并单元格"
        
        # 写入数据
        ws['A1'] = '销售情况'
        ws['B1'] = '数量'
        ws['C1'] = '金额'
        ws['A2'] = '产品A'
        ws['B2'] = 100
        ws['C2'] = 1000
        ws['A3'] = '产品B'
        ws['B3'] = 200
        ws['C3'] = 2000
        
        # 合并单元格
        ws.merge_cells('A1:A3')  # 合并A1到A3
        ws.merge_cells('B1:C1')  # 合并B1到C1
        
        wb.save(str(test_file))
        print(f"✅ 创建测试文件: {test_file}")
        
        # 测试合并单元格处理
        print(f"🔧 测试合并单元格处理...")
        excel_processor._unmerge_cells_with_fill(ws)
        
        # 检查结果
        print(f"📊 处理后的数据:")
        for row in ws.iter_rows(values_only=True):
            print(f"  {list(row)}")
        
        print(f"✅ 合并单元格处理完成")
        
    except Exception as e:
        print(f"❌ 合并单元格处理失败: {e}")
        import traceback
        traceback.print_exc()

async def main():
    """主测试函数"""
    print("🚀 开始增强Excel处理器测试")
    
    # 运行各项测试
    await test_enhanced_excel_processor()
    await test_schema_split_detection()
    await test_multi_level_header_detection()
    await test_merged_cells_processing()
    
    print("\n🎉 所有测试完成！")

if __name__ == "__main__":
    asyncio.run(main())
